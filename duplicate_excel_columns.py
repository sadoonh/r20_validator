#!/usr/bin/env python3
"""Build an Excel table with hidden control and variance columns."""

from __future__ import annotations

from copy import copy
from pathlib import Path
import re
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Protection
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from textual import on, work
from textual.app import App, ComposeResult
from textual.containers import Horizontal
from textual.screen import Screen
from textual.validation import Integer
from textual.widgets import (
    Button,
    DataTable,
    DirectoryTree,
    Footer,
    Header,
    Input,
    Label,
    OptionList,
    Static,
)

from excel_audit import install_audit_log


CELL_REFERENCE = re.compile(r"(?<![A-Z0-9_])(\$?)([A-Z]{1,3})(\$?\d+)")
TRUE_FILL = PatternFill(patternType="solid", fgColor="FFC6EFCE", bgColor="FFC6EFCE")
FALSE_FILL = PatternFill(patternType="solid", fgColor="FFFFC7CE", bgColor="FFFFC7CE")
CONTROL_FILL = PatternFill("solid", fgColor="FFD9D9D9")


def _target_column(column: int, table_start: int, table_end: int) -> int:
    if column < table_start:
        return column
    if column <= table_end:
        return table_start + ((column - table_start) * 3)
    return column + ((table_end - table_start + 1) * 2)


def _remap_formula(formula: str, table_start: int, table_end: int) -> str:
    """Point common A1 references at cells shifted by helper columns."""

    def replace_reference(match: re.Match[str]) -> str:
        column = column_index_from_string(match.group(2))
        new_column = get_column_letter(
            _target_column(column, table_start, table_end)
        )
        return f"{match.group(1)}{new_column}{match.group(3)}"

    return CELL_REFERENCE.sub(replace_reference, formula)


def _is_missing_value(value: object) -> bool:
    if value is None:
        return True
    return isinstance(value, str) and (
        value == "" or value.upper() == "#N/A"
    )


def _control_value(value: object) -> object:
    if _is_missing_value(value):
        return None
    if isinstance(value, str) and value.startswith("="):
        expression = value[1:]
        return f'=IF(ISNA({expression}),"",{expression})'
    return value


def _missing_formula(reference: str) -> str:
    return (
        f'OR(IFERROR({reference}="",FALSE),ISNA({reference}),'
        f'IFERROR({reference}="#N/A",FALSE))'
    )


def _variance_formula(original: str, control: str) -> str:
    return (
        f"=IF(AND({_missing_formula(original)},{_missing_formula(control)}),"
        f'"",IFERROR({original}={control},FALSE))'
    )


def _write_cell(target, value, style, hyperlink, comment) -> None:
    target.value = value
    target._style = copy(style)
    target.comment = copy(comment) if comment else None
    target._hyperlink = copy(hyperlink) if hyperlink else None
    if target._hyperlink:
        target._hyperlink.ref = target.coordinate


def _set_cell(target, value, style, hyperlink, comment, locked: bool) -> None:
    _write_cell(target, value, style, hyperlink, comment)
    target.protection = Protection(locked=locked, hidden=locked)


def _clear_cell(cell) -> None:
    cell.value = None
    cell._style = StyleArray()
    cell.comment = None
    cell._hyperlink = None


def _set_dimension(worksheet, column: int, source_dimension) -> None:
    letter = get_column_letter(column)
    dimension = copy(source_dimension)
    dimension.index = letter
    dimension.min = column
    dimension.max = column
    worksheet.column_dimensions[letter] = dimension


def _unique_table_name(workbook) -> str:
    existing_names = {
        table.displayName
        for sheet in workbook.worksheets
        for table in sheet.tables.values()
    }
    name = "ProcessedTable"
    suffix = 2
    while name in existing_names:
        name = f"ProcessedTable{suffix}"
        suffix += 1
    return name


def _remap_sort_state(sort_state, table_start: int, table_end: int) -> None:
    if sort_state is None:
        return
    if sort_state.ref:
        sort_state.ref = _remap_formula(
            sort_state.ref, table_start, table_end
        )
    for condition in sort_state.sortCondition:
        condition.ref = _remap_formula(
            condition.ref, table_start, table_end
        )


def _shift_unrelated_table(
    table: Table, table_start: int, table_end: int
) -> None:
    table.ref = _remap_formula(table.ref, table_start, table_end)
    if table.autoFilter:
        if table.autoFilter.ref:
            table.autoFilter.ref = _remap_formula(
                table.autoFilter.ref, table_start, table_end
            )
        _remap_sort_state(
            table.autoFilter.sortState, table_start, table_end
        )
    _remap_sort_state(table.sortState, table_start, table_end)


def duplicate_columns(
    input_path: Path, sheet_name: str | None, header_row: int = 1
) -> Path:
    workbook = load_workbook(input_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    if header_row < 1 or header_row > worksheet.max_row:
        raise ValueError(
            f"Header row must be between 1 and {worksheet.max_row}; got {header_row}"
        )

    source_max_column = worksheet.max_column
    source_max_row = worksheet.max_row
    existing_tables = list(worksheet.tables.values())
    matching_tables = [
        table
        for table in existing_tables
        if range_boundaries(table.ref)[1] == header_row
    ]
    if len(matching_tables) > 1:
        raise ValueError(
            "Multiple Excel tables start on the selected header row. "
            "Choose a row with a single table."
        )

    selected_table = matching_tables[0] if matching_tables else None
    if selected_table:
        table_start, _, table_end, table_bottom = range_boundaries(
            selected_table.ref
        )
        if table_bottom == header_row:
            raise ValueError("No data rows were found below the selected header.")
    else:
        populated_header_columns = [
            column
            for column in range(1, source_max_column + 1)
            if worksheet.cell(header_row, column).value not in (None, "")
        ]
        if not populated_header_columns:
            raise ValueError("The selected header row is empty.")
        table_start = populated_header_columns[0]
        table_end = populated_header_columns[-1]

        table_bottom = header_row
        for row in range(source_max_row, header_row, -1):
            if any(
                worksheet.cell(row, column).value not in (None, "")
                for column in range(table_start, table_end + 1)
            ):
                table_bottom = row
                break
        if table_bottom == header_row:
            raise ValueError("No data rows were found below the selected header.")

    unrelated_tables = [
        table for table in existing_tables if table is not selected_table
    ]
    for table in unrelated_tables:
        min_column, _, max_column, _ = range_boundaries(table.ref)
        if min_column <= table_end and max_column >= table_start:
            raise ValueError(
                f'Excel table "{table.displayName}" intersects the selected columns.'
            )

    source_headers: list[str] = []
    for column in range(table_start, table_end + 1):
        value = worksheet.cell(header_row, column).value
        if value in (None, ""):
            raise ValueError("The selected header row contains a blank column name.")
        source_headers.append(str(value))

    normalized_headers = [header.casefold() for header in source_headers]
    if len(normalized_headers) != len(set(normalized_headers)):
        raise ValueError("Excel table column names must be unique.")

    table_column_count = table_end - table_start + 1
    source_dimensions = {
        column: copy(worksheet.column_dimensions[get_column_letter(column)])
        for column in range(table_start, source_max_column + 1)
    }

    # Temporarily remove merges so their cells can move, then recreate each
    # merge over the corresponding shifted columns.
    merged_ranges = [
        range_boundaries(str(cell_range))
        for cell_range in worksheet.merged_cells.ranges
    ]
    for cell_range in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(cell_range))

    # Move all content from the table's first column onward. This preserves
    # titles and notes above the selected header while leaving blank slots for
    # control and variance columns.
    for source_column in range(source_max_column, table_start - 1, -1):
        target_column = _target_column(source_column, table_start, table_end)
        for row in range(1, source_max_row + 1):
            source = worksheet.cell(row, source_column)
            value = source.value
            style = copy(source._style)
            hyperlink = copy(source.hyperlink) if source.hyperlink else None
            comment = copy(source.comment) if source.comment else None
            if isinstance(value, str) and value.startswith("="):
                value = _remap_formula(value, table_start, table_end)

            target = worksheet.cell(row, target_column)
            _write_cell(target, value, style, hyperlink, comment)
            if target_column != source_column:
                _clear_cell(source)

    for letter in list(worksheet.column_dimensions):
        if column_index_from_string(letter) >= table_start:
            del worksheet.column_dimensions[letter]
    for source_column, source_dimension in source_dimensions.items():
        _set_dimension(
            worksheet,
            _target_column(source_column, table_start, table_end),
            source_dimension,
        )

    for min_column, min_row, max_column, max_row in merged_ranges:
        new_min_column = _target_column(min_column, table_start, table_end)
        new_max_column = _target_column(max_column, table_start, table_end)
        worksheet.merge_cells(
            start_row=min_row,
            start_column=new_min_column,
            end_row=max_row,
            end_column=new_max_column,
        )

    used_headers = set(normalized_headers)

    def unique_helper_header(base: str) -> str:
        candidate = base
        suffix = 2
        while candidate.casefold() in used_headers:
            candidate = f"{base}_{suffix}"
            suffix += 1
        used_headers.add(candidate.casefold())
        return candidate

    worksheet.sheet_properties.outlinePr.summaryRight = True
    worksheet.sheet_view.showOutlineSymbols = True
    variance_columns: list[int] = []
    helper_headers: list[tuple[str, str]] = []

    for index, source_header in enumerate(source_headers):
        original_column = table_start + (index * 3)
        control_column = original_column + 1
        variance_column = original_column + 2
        original_letter = get_column_letter(original_column)
        control_letter = get_column_letter(control_column)
        variance_letter = get_column_letter(variance_column)
        variance_columns.append(variance_column)
        source_dimension = source_dimensions[table_start + index]

        original_dimension = worksheet.column_dimensions[original_letter]
        original_dimension.protection = Protection(locked=False, hidden=False)
        for helper_column in (control_column, variance_column):
            _set_dimension(worksheet, helper_column, source_dimension)
            helper_dimension = worksheet.column_dimensions[
                get_column_letter(helper_column)
            ]
            helper_dimension.hidden = True
            helper_dimension.outlineLevel = 1
            helper_dimension.protection = Protection(locked=True, hidden=True)

        worksheet.column_dimensions[
            get_column_letter(variance_column + 1)
        ].collapsed = True

        control_header = unique_helper_header(f"{source_header}_control")
        variance_header = unique_helper_header(f"{source_header}_variance")
        helper_headers.append((control_header, variance_header))
        original_header_cell = worksheet.cell(header_row, original_column)
        original_header_cell.value = source_header
        original_header_cell.protection = Protection(locked=False, hidden=False)
        header_style = copy(original_header_cell._style)
        control_header_cell = worksheet.cell(header_row, control_column)
        _set_cell(
            control_header_cell,
            control_header,
            header_style,
            None,
            None,
            locked=True,
        )
        control_header_cell.fill = copy(CONTROL_FILL)
        _set_cell(
            worksheet.cell(header_row, variance_column),
            variance_header,
            header_style,
            None,
            None,
            locked=True,
        )

        for row in range(header_row + 1, table_bottom + 1):
            original = worksheet.cell(row, original_column)
            original.protection = Protection(locked=False, hidden=False)
            style = copy(original._style)
            control = worksheet.cell(row, control_column)
            variance = worksheet.cell(row, variance_column)
            _set_cell(
                control,
                _control_value(original.value),
                style,
                original.hyperlink,
                original.comment,
                locked=True,
            )
            control.fill = copy(CONTROL_FILL)
            _set_cell(
                variance,
                _variance_formula(
                    f"{original_letter}{row}", f"{control_letter}{row}"
                ),
                style,
                None,
                None,
                locked=True,
            )
            variance.number_format = "General"

        variance_range = (
            f"{variance_letter}{header_row + 1}:"
            f"{variance_letter}{table_bottom}"
        )
        variance_reference = f"${variance_letter}{header_row + 1}"
        worksheet.conditional_formatting.add(
            variance_range,
            FormulaRule(
                formula=[
                    f'IFERROR(AND({variance_reference}<>"",'
                    f"{variance_reference}=TRUE),FALSE)"
                ],
                fill=TRUE_FILL,
                stopIfTrue=True,
            ),
        )
        worksheet.conditional_formatting.add(
            variance_range,
            FormulaRule(
                formula=[
                    f'IFERROR(AND({variance_reference}<>"",'
                    f"{variance_reference}=FALSE),FALSE)"
                ],
                fill=FALSE_FILL,
                stopIfTrue=True,
            ),
        )

    # Add one visible, locked status column. Its filter can show every row that
    # contains at least one mismatch across the individual variance columns.
    overall_column = table_start + (table_column_count * 3)
    overall_letter = get_column_letter(overall_column)
    overall_dimension_source = source_dimensions[table_end]
    _set_dimension(worksheet, overall_column, overall_dimension_source)
    overall_dimension = worksheet.column_dimensions[overall_letter]
    overall_dimension.width = max(overall_dimension.width or 0, 18)
    overall_dimension.hidden = False
    overall_dimension.outlineLevel = 0
    overall_dimension.collapsed = True
    overall_dimension.protection = Protection(locked=False, hidden=False)

    overall_header = unique_helper_header("variance_filter")
    overall_header_style = copy(worksheet.cell(header_row, table_start)._style)
    _set_cell(
        worksheet.cell(header_row, overall_column),
        overall_header,
        overall_header_style,
        None,
        None,
        locked=False,
    )

    for row in range(header_row + 1, table_bottom + 1):
        variance_references = [
            f"{get_column_letter(column)}{row}" for column in variance_columns
        ]
        has_variance = "OR(" + ",".join(
            f'{reference}<>""' for reference in variance_references
        ) + ")"
        all_matching = "AND(" + ",".join(
            f'IF({reference}="",TRUE,{reference})'
            for reference in variance_references
        ) + ")"
        overall = worksheet.cell(row, overall_column)
        overall_style = copy(worksheet.cell(row, table_start)._style)
        _set_cell(
            overall,
            f'=IF({has_variance},{all_matching},"")',
            overall_style,
            None,
            None,
            locked=False,
        )
        overall.number_format = "General"

    overall_range = (
        f"{overall_letter}{header_row + 1}:{overall_letter}{table_bottom}"
    )
    overall_reference = f"${overall_letter}{header_row + 1}"
    worksheet.conditional_formatting.add(
        overall_range,
        FormulaRule(
            formula=[
                f'IFERROR(AND({overall_reference}<>"",'
                f"{overall_reference}=TRUE),FALSE)"
            ],
            fill=TRUE_FILL,
            stopIfTrue=True,
        ),
    )
    worksheet.conditional_formatting.add(
        overall_range,
        FormulaRule(
            formula=[
                f'IFERROR(AND({overall_reference}<>"",'
                f"{overall_reference}=FALSE),FALSE)"
            ],
            fill=FALSE_FILL,
            stopIfTrue=True,
        ),
    )

    table_right = overall_column
    table_ref = (
        f"{get_column_letter(table_start)}{header_row}:"
        f"{get_column_letter(table_right)}{table_bottom}"
    )
    if selected_table:
        original_columns = list(selected_table.tableColumns)
        if len(original_columns) != table_column_count:
            raise ValueError(
                f'Excel table "{selected_table.displayName}" has inconsistent '
                "column metadata."
            )

        rebuilt_columns: list[TableColumn] = []
        for index, (source_header, helper_names) in enumerate(
            zip(source_headers, helper_headers)
        ):
            original_column = copy(original_columns[index])
            original_column.id = len(rebuilt_columns) + 1
            original_column.name = source_header
            rebuilt_columns.append(original_column)
            for helper_name in helper_names:
                rebuilt_columns.append(
                    TableColumn(
                        id=len(rebuilt_columns) + 1,
                        name=helper_name,
                    )
                )
        rebuilt_columns.append(
            TableColumn(id=len(rebuilt_columns) + 1, name=overall_header)
        )

        selected_table.ref = table_ref
        selected_table.tableColumns = rebuilt_columns
        if selected_table.autoFilter:
            if selected_table.autoFilter.ref:
                _, filter_top, _, filter_bottom = range_boundaries(
                    selected_table.autoFilter.ref
                )
            else:
                filter_top, filter_bottom = header_row, table_bottom
            selected_table.autoFilter.ref = (
                f"{get_column_letter(table_start)}{filter_top}:"
                f"{get_column_letter(table_right)}{filter_bottom}"
            )
            for filter_column in selected_table.autoFilter.filterColumn:
                if 0 <= filter_column.colId < table_column_count:
                    filter_column.colId *= 3
            _remap_sort_state(
                selected_table.autoFilter.sortState,
                table_start,
                table_end,
            )
        _remap_sort_state(selected_table.sortState, table_start, table_end)
    else:
        table = Table(displayName=_unique_table_name(workbook), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    for table in unrelated_tables:
        _shift_unrelated_table(table, table_start, table_end)

    # Sheet protection is required for locked helper cells. Formatting columns
    # remains allowed so users can expand/collapse the outline groups.
    protection = worksheet.protection
    protection.sheet = True
    protection.formatColumns = False
    protection.formatCells = False
    protection.formatRows = False
    protection.insertRows = False
    protection.deleteRows = False
    protection.sort = False
    protection.autoFilter = False
    protection.selectUnlockedCells = False

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_path = input_path.with_name(
        f"{input_path.stem}_processed{input_path.suffix}"
    )
    workbook.save(output_path)
    return output_path


class ExcelDirectoryTree(DirectoryTree):
    """Directory tree that only displays folders and Excel workbooks."""

    def filter_paths(self, paths: Iterable[Path]) -> Iterable[Path]:
        return [
            path
            for path in paths
            if path.is_dir()
            or (path.suffix.lower() == ".xlsx" and not path.name.startswith("~$"))
        ]


class FilePickerScreen(Screen[None]):
    BINDINGS = [("q", "app.quit", "Quit")]

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        yield Static("1. Choose an Excel file", classes="screen-title")
        yield Static(
            f"Showing .xlsx files under {Path.cwd()}", classes="screen-help"
        )
        yield ExcelDirectoryTree(Path.cwd(), id="file-tree")
        yield Footer()

    @on(DirectoryTree.FileSelected, "#file-tree")
    def select_file(self, event: DirectoryTree.FileSelected) -> None:
        event.stop()
        self.app.open_workbook(event.path)


class SheetPickerScreen(Screen[None]):
    BINDINGS = [("escape", "app.pop_screen", "Back"), ("q", "app.quit", "Quit")]

    def __init__(self, file_path: Path, sheet_names: list[str]) -> None:
        super().__init__()
        self.file_path = file_path
        self.sheet_names = sheet_names

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        yield Static("2. Choose a worksheet", classes="screen-title")
        yield Static(str(self.file_path), classes="screen-help")
        yield OptionList(*self.sheet_names, id="sheet-list")
        yield Footer()

    @on(OptionList.OptionSelected, "#sheet-list")
    def select_sheet(self, event: OptionList.OptionSelected) -> None:
        self.app.open_sheet(
            self.file_path, self.sheet_names[event.option_index]
        )


class HeaderPickerScreen(Screen[None]):
    BINDINGS = [("escape", "app.pop_screen", "Back"), ("q", "app.quit", "Quit")]

    def __init__(self, file_path: Path, sheet_name: str, raw_data: pd.DataFrame) -> None:
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.raw_data = raw_data

    def compose(self) -> ComposeResult:
        row_count = len(self.raw_data.index)
        yield Header(show_clock=True)
        yield Static("3. Choose the header row", classes="screen-title")
        yield Static(
            f"{self.file_path.name}  •  {self.sheet_name}  •  "
            f"{row_count} detected rows\n"
            "df.head(10): select a row; its first value anchors the preserved sheet's table.",
            classes="screen-help",
        )
        yield DataTable(id="preview", cursor_type="row", zebra_stripes=True)
        with Horizontal(id="header-controls"):
            yield Label("Header row:")
            yield Input(
                placeholder=f"1–{row_count}",
                type="integer",
                validators=Integer(minimum=1, maximum=row_count),
                id="header-row",
            )
            yield Button("Process workbook", variant="primary", id="process")
        yield Static("", id="status")
        yield Footer()

    def on_mount(self) -> None:
        table = self.query_one("#preview", DataTable)
        preview = self.raw_data.head(10)
        table.add_column("Excel row", key="excel-row")
        for column_number in range(1, len(preview.columns) + 1):
            table.add_column(get_column_letter(column_number))

        for index, row in preview.iterrows():
            values = [_display_value(value) for value in row.tolist()]
            table.add_row(str(index + 1), *values, key=str(index + 1))
        table.focus()

    @on(DataTable.RowSelected, "#preview")
    def use_selected_row(self, event: DataTable.RowSelected) -> None:
        self.query_one("#header-row", Input).value = str(event.cursor_row + 1)

    @on(Input.Submitted, "#header-row")
    def submit_header_row(self) -> None:
        self.start_processing()

    @on(Button.Pressed, "#process")
    def process_button_pressed(self) -> None:
        self.start_processing()

    def start_processing(self) -> None:
        input_widget = self.query_one("#header-row", Input)
        try:
            header_row = int(input_widget.value)
        except ValueError:
            self.notify("Enter a valid row number.", severity="error")
            return

        if not 1 <= header_row <= len(self.raw_data.index):
            self.notify(
                f"Header row must be between 1 and {len(self.raw_data.index)}.",
                severity="error",
            )
            return

        self.query_one("#process", Button).disabled = True
        self.query_one("#status", Static).update("Processing workbook…")
        self.process_workbook(header_row)

    @work(thread=True, exclusive=True)
    def process_workbook(self, header_row: int) -> None:
        try:
            output_path = duplicate_columns(
                self.file_path, self.sheet_name, header_row
            )
            output_path = install_audit_log(output_path)
        except Exception as error:
            self.app.call_from_thread(self.show_error, str(error))
            return
        self.app.call_from_thread(self.show_result, output_path)

    def show_error(self, message: str) -> None:
        self.query_one("#process", Button).disabled = False
        self.query_one("#status", Static).update("")
        self.notify(message, title="Processing failed", severity="error")

    def show_result(self, output_path: Path) -> None:
        self.app.push_screen(CompleteScreen(output_path))


class CompleteScreen(Screen[None]):
    BINDINGS = [("q", "app.quit", "Quit")]

    def __init__(self, output_path: Path) -> None:
        super().__init__()
        self.output_path = output_path

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        yield Static("Workbook processed", classes="screen-title success")
        yield Static(
            f"Created:\n{self.output_path}", id="complete-message"
        )
        yield Button("Quit", variant="primary", id="quit")
        yield Footer()

    @on(Button.Pressed, "#quit")
    def quit_app(self) -> None:
        self.app.exit()


def _display_value(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


class ExcelProcessorApp(App[None]):
    TITLE = "Excel Control Column Prototype"
    CSS = """
    Screen {
        layout: vertical;
    }

    .screen-title {
        height: 2;
        padding: 0 1;
        content-align: left middle;
        text-style: bold;
        color: $accent;
    }

    .screen-help {
        height: 2;
        padding: 0 1;
        color: $text-muted;
        overflow: hidden;
    }

    #file-tree, #sheet-list, #preview {
        height: 1fr;
        margin: 0 1;
        border: round $accent;
    }

    #header-controls {
        height: 3;
        padding: 0 1;
        align: left middle;
    }

    #header-controls Label {
        width: 13;
        padding-top: 1;
    }

    #header-row {
        width: 18;
        margin-right: 2;
    }

    #process {
        width: 22;
    }

    #status {
        height: 1;
        padding: 0 1;
        color: $warning;
    }

    #complete-message {
        width: 80%;
        height: auto;
        margin: 2 4;
        padding: 2;
        border: round $success;
    }

    CompleteScreen Button {
        width: 20;
        margin: 1 4;
    }

    .success {
        color: $success;
    }
    """

    def on_mount(self) -> None:
        self.push_screen(FilePickerScreen())

    def open_workbook(self, file_path: Path) -> None:
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
        except Exception as error:
            self.notify(str(error), title="Could not open workbook", severity="error")
            return
        self.push_screen(SheetPickerScreen(file_path, sheet_names))

    def open_sheet(self, file_path: Path, sheet_name: str) -> None:
        try:
            raw_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        except Exception as error:
            self.notify(str(error), title="Could not read worksheet", severity="error")
            return

        if raw_data.empty:
            self.notify("The selected worksheet is empty.", severity="error")
            return
        self.push_screen(HeaderPickerScreen(file_path, sheet_name, raw_data))


def main() -> None:
    ExcelProcessorApp().run()


if __name__ == "__main__":
    main()
