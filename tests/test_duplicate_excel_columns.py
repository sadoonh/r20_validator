from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.filters import FilterColumn, Filters
from openpyxl.worksheet.table import Table, TableStyleInfo

from duplicate_excel_columns import duplicate_columns


class ExistingTableTests(unittest.TestCase):
    def test_rebuilds_existing_table_and_preserves_its_identity(self) -> None:
        with TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Name", "Amount"])
            worksheet.append(["Alpha", 10])
            worksheet.append(["Beta", 20])

            table = Table(displayName="ExistingSales", ref="A1:B3")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showRowStripes=True
            )
            table._initialise_columns()
            table.tableColumns[0].name = "Name"
            table.tableColumns[1].name = "Amount"
            table.autoFilter.filterColumn.append(
                FilterColumn(colId=1, filters=Filters(filter=["10"]))
            )
            worksheet.add_table(table)
            workbook.save(input_path)

            output_path = duplicate_columns(input_path, None, header_row=1)
            result = load_workbook(output_path, data_only=False)
            result_sheet = result.active
            rebuilt = result_sheet.tables["ExistingSales"]

            self.assertEqual(rebuilt.ref, "A1:G3")
            self.assertEqual(rebuilt.tableStyleInfo.name, "TableStyleMedium9")
            self.assertEqual(
                [column.name for column in rebuilt.tableColumns],
                [
                    "Name",
                    "Name_control",
                    "Name_variance",
                    "Amount",
                    "Amount_control",
                    "Amount_variance",
                    "variance_filter",
                ],
            )
            self.assertEqual(rebuilt.autoFilter.filterColumn[0].colId, 3)
            for coordinate in ("B1", "B2", "E1", "E2"):
                self.assertEqual(result_sheet[coordinate].fill.fill_type, "solid")
                self.assertEqual(
                    result_sheet[coordinate].fill.fgColor.rgb, "FFD9D9D9"
                )

            conditional_rules = {
                str(formatting.sqref): formatting.rules
                for formatting in result_sheet.conditional_formatting
            }
            variance_rules = conditional_rules["C2:C3"]
            self.assertIn("$A2=$B2", variance_rules[0].formula[0])
            self.assertIn("$A2=$B2", variance_rules[1].formula[0])
            self.assertEqual(variance_rules[0].dxf.fill.fgColor.rgb, "FFC6EFCE")
            self.assertEqual(variance_rules[1].dxf.fill.fgColor.rgb, "FFFFC7CE")
            overall_rules = conditional_rules["G2:G3"]
            self.assertIn("$A2=$B2", overall_rules[0].formula[0])
            self.assertIn("$D2=$E2", overall_rules[0].formula[0])

            self.assertIn("ISNA(A2)", result_sheet["C2"].value)
            self.assertIn("IFERROR(A2=B2,FALSE)", result_sheet["C2"].value)
            self.assertEqual(
                result_sheet["G2"].value,
                '=IF(OR(C2<>"",F2<>""),'
                'AND(IF(C2="",TRUE,C2),IF(F2="",TRUE,F2)),"")',
            )

    def test_missing_values_produce_blank_controls_and_variances(self) -> None:
        with TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Blank", "Error", "FormulaError", "FormulaBlank"])
            worksheet.append([None, "#N/A", "=NA()", '=""'])
            worksheet.add_table(Table(displayName="MissingValues", ref="A1:D2"))
            workbook.save(input_path)

            output_path = duplicate_columns(input_path, None, header_row=1)
            result_sheet = load_workbook(output_path, data_only=False).active

            self.assertIsNone(result_sheet["B2"].value)
            self.assertIsNone(result_sheet["E2"].value)
            self.assertEqual(result_sheet["H2"].value, '=IFNA(NA(),"")')
            self.assertEqual(result_sheet["K2"].value, '=IFNA("","")')
            for coordinate in ("C2", "F2", "I2", "L2"):
                formula = result_sheet[coordinate].value
                self.assertTrue(formula.startswith("=IF(AND("))
                self.assertIn('),"",IFERROR(', formula)

            self.assertEqual(
                result_sheet["M2"].value,
                '=IF(OR(C2<>"",F2<>"",I2<>"",L2<>""),'
                'AND(IF(C2="",TRUE,C2),IF(F2="",TRUE,F2),'
                'IF(I2="",TRUE,I2),IF(L2="",TRUE,L2)),"")',
            )

    def test_shifts_an_unrelated_table_to_the_right(self) -> None:
        with TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            for row in (("A", "B"), (1, 2), (3, 4)):
                worksheet.append(row)
            worksheet["J5"] = "X"
            worksheet["K5"] = "Y"
            worksheet["J6"] = 5
            worksheet["K6"] = 6
            worksheet.add_table(Table(displayName="Main", ref="A1:B3"))
            worksheet.add_table(Table(displayName="Other", ref="J5:K6"))
            workbook.save(input_path)

            output_path = duplicate_columns(input_path, None, header_row=1)
            result_sheet = load_workbook(output_path).active

            self.assertEqual(result_sheet.tables["Main"].ref, "A1:G3")
            self.assertEqual(result_sheet.tables["Other"].ref, "N5:O6")
            self.assertEqual(result_sheet["N6"].value, 5)
            self.assertEqual(result_sheet["O6"].value, 6)


if __name__ == "__main__":
    unittest.main()
